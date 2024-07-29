import sqlite3
from datetime import datetime, timedelta
import openpyxl
import time

conn = sqlite3.connect('/usr/share/hassio/homeassistant/home-assistant_v2.db')
cursor = conn.cursor()

metadata_ids = [15, 16, 12, 13, 17, 18, 19, 42, 43, 44, 45, 46, 54]

csv_file = './test.xlsx'

workbook = openpyxl.load_workbook(csv_file)
sheet = workbook.active

last_row = sheet.max_row

new_data = []

row_num = 5

def main():
    global row_num
    new_data.clear()
    current_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for metadata_id in metadata_ids:
        query = f"SELECT * FROM statistics_short_term WHERE metadata_id = {metadata_id} limit 1"
        cursor.execute(query)
        rows = cursor.fetchall()

        if rows:
            for row in rows:
                data = {
                    "日時": current_timestamp,
                    "電力積算値": row[7]
                }
                new_data.append(data)

    sheet.cell(row=row_num, column=1, value=new_data[1]['日時'])

    sheet.cell(row=row_num, column=2, value=new_data[0]['電力積算値'])
    sheet.cell(row=row_num, column=3, value=new_data[1]['電力積算値'])
    sheet.cell(row=row_num, column=4, value=new_data[2]['電力積算値'])
    sheet.cell(row=row_num, column=5, value=new_data[3]['電力積算値'])
    sheet.cell(row=row_num, column=6, value=new_data[4]['電力積算値'])
    sheet.cell(row=row_num, column=7, value=new_data[5]['電力積算値'])
    sheet.cell(row=row_num, column=8, value=new_data[6]['電力積算値'])
    sheet.cell(row=row_num, column=9, value=new_data[7]['電力積算値'])
    sheet.cell(row=row_num, column=10, value=new_data[8]['電力積算値'])
    sheet.cell(row=row_num, column=11, value=new_data[9]['電力積算値'])
    sheet.cell(row=row_num, column=12, value=new_data[10]['電力積算値'])
    sheet.cell(row=row_num, column=13, value=new_data[11]['電力積算値'])
    sheet.cell(row=row_num, column=14, value=new_data[12]['電力積算値'])

    workbook.save(csv_file)
    row_num += 1

if __name__ == '__main__':
    try:
        while True:
            now = datetime.now()
            if now.minute == 22 or now.minute == 23:
                main()
            next_minute = now.replace(second=0, microsecond=0) + timedelta(minutes=1)
            sleep_time = (next_minute - now).total_seconds()
            time.sleep(sleep_time)
    except KeyboardInterrupt:
        conn.close()
        workbook.save(csv_file)
